'''
Created on Nov 12, 2016
'''
from openpyxl import load_workbook, Workbook
from sys import exit
import os
from Tkinter import Tk
from tkFileDialog import askdirectory
import logging.handlers
from win32com import client
from datetime import datetime
from copy import copy

EMAIL_FILE = 'emails.txt'
SUBJECT_FILE = 'subject.txt'
BODY_FILE = 'body.txt'
MAX_LOG_SIZE = 2000000
LOG_FILENAME = 'sheetEmailer.log'
BACKUP_LOGS = 5
CURRENT_DATETIME = datetime.now().strftime('%Y-%m-%d-%H-%M-%S')

class SheetEmailer:

    def __init__(self):
        self.absolute_file_paths = []
        self.client_to_email = dict()
        self.all_client_numbers = set()
        self.excel = None
        self.outlook = None
        self.subject = ''
        self.body = ''
        self._read_emails()
        self._read_subject()
        self._read_body()

        logger.debug('Please choose a directory')
        Tk().withdraw() # we don't want a full GUI, so keep the root window from appearing
        directory = askdirectory()

        if directory:
            logger.debug('Chosen directory: ' + directory)
            self.workbook_directory = directory
        else:
            exit('Please choose a valid directory')

    def _read_emails(self):
        email_sources_path = os.path.abspath(EMAIL_FILE)
        logger.debug('Loading email address from ' + email_sources_path)
        try:
            email_file = open(EMAIL_FILE, 'r')
        except:
            exit(EMAIL_FILE + ' does not exist')

        for line in email_file:
            email_mapping = line.split('=')
            client_number = email_mapping[0]
            client_email = email_mapping[1]
            self.client_to_email[client_number] = client_email.strip().split(',')
            self.client_to_email[client_number] = [c.strip() for c in self.client_to_email[client_number]]

            logger.debug('Email mapping ' + str(self.client_to_email))

    def _read_subject(self):
        subject_sources_path = os.path.abspath(SUBJECT_FILE)
        logger.debug('Loading subject from ' + subject_sources_path)
        try:
            with open(subject_sources_path, 'r') as content_file:
                self.subject = content_file.read()
        except:
            exit(subject_sources_path + ' does not exist')

        logger.debug('Subject: ' + self.subject)

    def _read_body(self):
        body_sources_path = os.path.abspath(BODY_FILE)
        logger.debug('Loading body from ' + body_sources_path)
        try:
            with open(body_sources_path, 'r') as content_file:
                self.body = content_file.read()
        except:
            exit(body_sources_path + ' does not exist')

        logger.debug('Body: ' + self.body)

    def create_user_directories_and_files(self):
        for filename in os.listdir(self.workbook_directory):
            logger.debug('Working on workbook: ' + filename)
            workbook_name = filename.split('.')[0]
            full_workbook_path = os.path.join(self.workbook_directory, filename)
            try:
                loaded_workbook = load_workbook(filename = full_workbook_path)
            except:
                logger.debug("Unable to load workbook " + filename + ". Will continue to next workbook.")
                continue


            for sheet_name in loaded_workbook.sheetnames:
                logger.debug('Working on sheet: ' + sheet_name)
                current_sheet = loaded_workbook[sheet_name]

                new_workbook, user_number = self._create_workbook_from_sheet(current_sheet)
                self.all_client_numbers.add(user_number)
                self._save_workbook(user_number, workbook_name, new_workbook)

    def _create_workbook_from_sheet(self, current_sheet):
        new_workbook = Workbook()
        active_sheet = new_workbook.active
        current_sheet_rows = current_sheet.rows

        user_number = current_sheet.oddFooter.left.text.strip()

        for row_num, row in enumerate(current_sheet_rows):
            for col_num, cell in enumerate(row):
                active_sheet_cell = active_sheet.cell(column=col_num + 1, row=row_num + 1)
                active_sheet_cell.value = cell.value
                active_sheet_cell.style = cell.style
                active_sheet_cell.font = copy(cell.font)
                active_sheet_cell.fill = copy(cell.fill)
                active_sheet_cell.border = copy(cell.border)
                active_sheet_cell.alignment = copy(cell.alignment)

        active_sheet.oddHeader.center = current_sheet.oddHeader.center
        active_sheet.evenHeader.center = current_sheet.evenHeader.center
        active_sheet.orientation= current_sheet.orientation

        return new_workbook, user_number

    def _save_workbook(self, user_number, workbook_name, new_workbook):
        path_to_new_workbook_directory = self._create_directory_if_none_exists(user_number)

        new_file_path = os.path.join(path_to_new_workbook_directory, workbook_name + '_' + user_number + '.xlsx')

        absolute_path_to_file = os.path.abspath(new_file_path)

        logger.debug("Writing " + absolute_path_to_file)

        self.absolute_file_paths.append(absolute_path_to_file)

        new_workbook.save(new_file_path)


    def _create_directory_if_none_exists(self, user_number):
        absoulte_path_to_sheet = os.path.abspath(os.path.join(CURRENT_DATETIME, user_number))
        if not os.path.exists(absoulte_path_to_sheet):
            logger.debug("Creating directory: " + absoulte_path_to_sheet)
            os.makedirs(absoulte_path_to_sheet)
        else:
            logger.debug("Directory already exists: " + absoulte_path_to_sheet)

        return absoulte_path_to_sheet

    def convert_all_to_pdf(self):
        logger.debug('Starting excel')
        self.excel = client.Dispatch("Excel.Application")
        for path in self.absolute_file_paths:
            logger.debug('Converting ' + path + ' to pdf')
            books = self.excel.Workbooks.Open(path)
            ws = books.Worksheets[0]
            ws.Visible = 1
            pdf_path = os.path.join(os.path.dirname(path), os.path.basename(path).split('.')[0] + '.pdf')
            ws.ExportAsFixedFormat(0, pdf_path)
            logger.debug('Exported ' + pdf_path)
        self.excel.quit()
        self.excel = None
        logger.debug('Quit Excel')


    def save_email_to_draft(self):
        logger.debug('Starting outlook')
        self.outlook = client.Dispatch('outlook.application')
        for user in self.all_client_numbers:
            absoulte_path_to_directory = os.path.abspath(os.path.join(CURRENT_DATETIME, user))
            if user in self.client_to_email:
                mail = self.outlook.CreateItem(0)
                mail.To = ','.join(self.client_to_email[user])
                mail.Subject = self.subject
                mail.HtmlBody = self.body
                for filename in os.listdir(absoulte_path_to_directory):
                    if filename.endswith(".pdf"):
                        path_to_pdf = os.path.join(absoulte_path_to_directory, filename)
                        logger.debug('Adding pdf: ' + path_to_pdf)
                        mail.Attachments.Add(Source=path_to_pdf )
                mail.Save()
            else:
                absoulte_path_to_no_email_directory = self._create_directory_if_none_exists('no_email')
                for filename in os.listdir(absoulte_path_to_directory):
                    os.rename(os.path.join(absoulte_path_to_directory, filename), os.path.join(absoulte_path_to_no_email_directory, filename))
                os.rmdir(absoulte_path_to_directory)
        self.outlook.quit()
        self.outlook = None
        logger.debug('Quit Outlook')

if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG,
                        format='%(asctime)s %(name)-12s %(levelname)-8s %(message)s',
                        datefmt='%m-%d %H:%M',
                        filename=LOG_FILENAME)

    console = logging.StreamHandler()
    console.setLevel(logging.DEBUG)
    formatter = logging.Formatter('%(name)-12s: %(levelname)-8s %(message)s')
    console.setFormatter(formatter)
    logging.getLogger('').addHandler(console)
    logger = logging.getLogger(__name__)
    handler = logging.handlers.RotatingFileHandler(
        LOG_FILENAME, maxBytes=MAX_LOG_SIZE, backupCount=BACKUP_LOGS)

    logger.addHandler(handler)


    logger.debug('Starting')
    try:
        sheet_emailer = SheetEmailer()
        sheet_emailer.create_user_directories_and_files()
        sheet_emailer.convert_all_to_pdf()
        sheet_emailer.save_email_to_draft()
    except Exception as e:
        if sheet_emailer.excel:
            sheet_emailer.excel.quit()
            sheet_emailer.excel = None
            logger.debug('Quit Excel in Except')

        if sheet_emailer.outlook:
            sheet_emailer.outlook.quit()
            sheet_emailer.outlook = None
            logger.debug('Quit Outlook in Except')

        logger.debug(e)

    logger.debug('Done')